from config import TOKEN
import logging
from aiogram import Bot, Dispatcher, executor, types
from aiogram.types import ContentType
import keyboards as kb
import sys
import time
import snap7
from Fun import get_real
import struct
import re
import os

import openpyxl
from openpyxl import workbook

help_text = """/list - просто вывод списка тегов. после команды можно ввести часть названия тега. 
/set - операция присваивания. Требуется что бы пароль был введен. set pr1_cd{ENTER} вывод значения тега и потом ввод нужного значения.
/pass - для операций присвоения тега потребуется ввод пароля
"/" - со слеша начинается описание части тега которого нужно просто вывести в консоли"""
# Захерачить сюда Zabbix мониторинг

try:
    wb = openpyxl.load_workbook(filename = 'Export.xlsx', read_only=True)
except:
    print('Проверте файл!!!')
    input('Нажмите ENTER.')
    sys.exit()

DB = {}
Connect_Controller = False
Access = False
count_tag = 0
ws1 = wb['Connections']
ws2 = wb['Groups']
ws3 = wb['Tags']

Connections = ws1['a1':'d'+str(ws1.max_row)]

Tag = ws3['a1':'g'+str(ws3.max_row)]

print("Начало работы бота.")

bot = Bot(token=TOKEN)
dp = Dispatcher(bot)

@dp.message_handler(commands=['start'])
async def process_hello(message: types.Message):
      await bot.send_message(message.from_user.id, 'Привет\nЧто будем делать?',
                             reply_markup=kb.greet_kb)

@dp.message_handler(commands=['Group'])
async def process_help(message: types.Message):
      await bot.send_message(message.from_user.id, 'Выберите уровень какой бочки вам показать?', reply_markup=kb.markupGroup)

@dp.message_handler(commands=['R1_R8'])
async def process_help(message: types.Message):
      await bot.send_message(message.from_user.id, 'Выберите бочку', reply_markup=kb.markupR)

@dp.message_handler(commands=['help'])
async def process_help(message: types.Message):
      await bot.send_message(message.from_user.id, help_text)

@dp.message_handler(commands=['rst'])
async def restart(event):
      python = sys.executable
      os.execl(python, python, * sys.argv)

@dp.message_handler(content_types=ContentType.ANY)
async def send_massage(message: types.Message):
      command = message.text.lower().split(' ')
      Connect_Controller = False 

      if message.text.lower() == 'привет':
            print ('Получил "привет"')
            await bot.send_message(message.chat.id, 'Привет, мой создатель')
      if message.text.lower() == 'пока':
            await bot.send_message(message.chat.id, 'Прощай, создатель')

      if message.text.lower()[0] == 'exit' or message.text.lower()[0] == '':
            Connect_Controller = False
            #client.disconnect()
            sys.exit()
      if message.text.lower()[0] == 'test' and len(message.text) == 2:
            print ('переход по 2 словам в команде')
            if message.text.lower()[0] == 'list':
                  print ('переход по LIST')
                  for i in range(3,ws3.max_row):
                        if Tag[i][0].value.lower().find(message.text.lower()[1]) >= 0:
                              await bot.send_message(message.chat.id,  Tag[i][0].value)
      if len(command) == 2:
            if command[0] == 'list':
                  for i in range(3,ws3.max_row):
                        if Tag[i][0].value.lower().find(command[1]) >= 0:
                              await bot.send_message(message.chat.id,  Tag[i][0].value)
      if (command[0][:1] == '/') or (command[0][:1] == '#'):                              # / - сделано для того что бы можно было повторно вызвать команду нажав на ссылку в которую преобразуется строка начинающаяся на /
            for i in range(3,ws3.max_row):
                  #print(command[0][1:])
                  if Tag[i][0].value.lower().find(command[0][1:]) >= 0:
                        for j in range(3,ws1.max_row):
                              if Tag[i][4].value == Connections[j][0].value:              # Если имя соединения правильное
                                    if Tag[i][6].value[:2] == 'DB':                         # Если в названии присутствует ДБ то значит мы сможем его прочитать из ДБ.
                                          tDB, tDD = re.findall('(\d+)', Tag[i][6].value)     # пишем во временные переменные номер ДБ и номер ДД
                                          if not Connect_Controller:
                                                client = snap7.client.Client()
                                                stat = client.connect(Connections[j][3].value.split(',')[1], int(Connections[j][3].value.split(',')[3]), int(Connections[j][3].value.split(',')[4]))
                                                Connect_Controller = True
                                          if Connect_Controller:
                                                DB[int(tDB)]     = client.db_read(int(tDB), 0, int(tDD)+4)
                                                Data        = get_real(DB[int(tDB)], int(tDD))
                                                send_data = Tag[i][0].value + '  -->  ' + str(round(Data, 2))
                                                print (Tag[i][0].value, '\t\t', Data)
                                                await bot.send_message(message.from_user.id, send_data)
            client.disconnect()
            Connect_Controller = False




if __name__ == '__main__':
    executor.start_polling(dp, skip_updates=True)